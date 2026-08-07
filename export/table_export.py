import csv
import os

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
)
from openpyxl.utils import get_column_letter


def prepare_rows(results):
    """
    Wandelt die internen Analyseergebnisse
    in einfache Tabellenzeilen um.
    """

    rows = []

    for result in results:

        filter_mode = result.get(
            "filter_mode",
            ""
        )

        if filter_mode == "light":
            filter_name = "Leicht"

        elif filter_mode == "medium":
            filter_name = "Mittel"

        else:
            filter_name = filter_mode

        rows.append(
            {
                "Bereich":
                    result["region_id"],

                "Gültige Pixel":
                    result["total_pixels"],

                "Grüne Pixel":
                    result["green_pixels"],

                "Grünanteil [%]":
                    round(
                        result["green_percentage"],
                        2
                    ),

                "Nicht Grün Pixel":
                    result["non_green_pixels"],

                "Nicht Grün [%]":
                    round(
                        result["non_green_percentage"],
                        2
                    ),

                "Grünfilter":
                    filter_name,
            }
        )

    return rows


def save_results_csv(
    results,
    file_path
):
    """
    Speichert Ergebnisse als CSV.
    """

    if not results:
        raise ValueError(
            "Keine Analyseergebnisse vorhanden."
        )

    rows = prepare_rows(
        results
    )

    fieldnames = list(
        rows[0].keys()
    )

    with open(
        file_path,
        "w",
        newline="",
        encoding="utf-8-sig"
    ) as csv_file:

        writer = csv.DictWriter(
            csv_file,
            fieldnames=fieldnames,
            delimiter=";"
        )

        writer.writeheader()

        writer.writerows(
            rows
        )

    return True


def save_results_excel(
    results,
    file_path
):
    """
    Speichert Ergebnisse als Excel-XLSX.
    """

    if not results:
        raise ValueError(
            "Keine Analyseergebnisse vorhanden."
        )

    rows = prepare_rows(
        results
    )

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = (
        "Grünanalyse"
    )

    headers = list(
        rows[0].keys()
    )

    # Überschriften schreiben
    for column_number, header in enumerate(
        headers,
        start=1
    ):

        cell = worksheet.cell(
            row=1,
            column=column_number,
            value=header
        )

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAD3"
        )

    # Daten schreiben
    for row_number, row_data in enumerate(
        rows,
        start=2
    ):

        for column_number, header in enumerate(
            headers,
            start=1
        ):

            value = row_data[
                header
            ]

            cell = worksheet.cell(
                row=row_number,
                column=column_number,
                value=value
            )

            if header in (
                "Grünanteil [%]",
                "Nicht Grün [%]",
            ):
                cell.number_format = (
                    '0.00'
                )

    # Spaltenbreiten automatisch einstellen
    for column_number, header in enumerate(
        headers,
        start=1
    ):

        max_length = len(
            str(header)
        )

        for row_number in range(
            2,
            worksheet.max_row + 1
        ):

            value = worksheet.cell(
                row=row_number,
                column=column_number
            ).value

            if value is not None:

                max_length = max(
                    max_length,
                    len(str(value))
                )

        worksheet.column_dimensions[
            get_column_letter(
                column_number
            )
        ].width = (
            max_length + 3
        )

    # Kopfzeile fixieren
    worksheet.freeze_panes = (
        "A2"
    )

    # Autofilter
    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    workbook.save(
        file_path
    )

    return True


def save_results(
    results,
    file_path
):
    """
    Erkennt automatisch anhand der Dateiendung,
    ob CSV oder XLSX gespeichert werden soll.
    """

    extension = os.path.splitext(
        file_path
    )[1].lower()

    if extension == ".csv":

        return save_results_csv(
            results,
            file_path
        )

    if extension == ".xlsx":

        return save_results_excel(
            results,
            file_path
        )

    raise ValueError(
        "Unterstützte Tabellenformate: CSV und XLSX."
    )